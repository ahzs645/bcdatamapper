#!/usr/bin/env python3
import csv
import json
import re
import sys
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from tempfile import TemporaryDirectory
from urllib.request import Request, urlopen

from openpyxl import load_workbook

OUTPUT_DIR = Path("public/data/wars")
SOURCE_PAGE = "https://www2.gov.bc.ca/gov/content/transportation/transportation-infrastructure/engineering-standards-guidelines/environmental-management/wildlife-management/wildlife-accident-reporting-system"
SOURCE_LICENSE = "WARS Data Use Licence Agreement 2026"
SOURCE_CITATION = "British Columbia Ministry of Transportation and Transit Wildlife Accident Reporting System, © 2025 Province of British Columbia. All rights reserved."

SOURCE_FILES = [
    {
        "id": "2018_2025",
        "title": "WARS 2018 to 2025",
        "url": "https://www2.gov.bc.ca/assets/gov/driving-and-transportation/transportation-infrastructure/engineering-standards-and-guidelines/environment/wars/historical-data/2018_to_2025_wars_historic_data.xlsx",
    },
    {
        "id": "2008_2018",
        "title": "WARS 2008 to 2018",
        "url": "https://www2.gov.bc.ca/assets/gov/driving-and-transportation/transportation-infrastructure/engineering-standards-and-guidelines/environment/wars/historical-data/2008_to_2018_wars_historic_data.xlsx",
    },
    {
        "id": "1998_2008",
        "title": "WARS 1998 to 2008",
        "url": "https://www2.gov.bc.ca/assets/gov/driving-and-transportation/transportation-infrastructure/engineering-standards-and-guidelines/environment/wars/historical-data/1998_to_2008_wars_historic_data.xlsx",
    },
    {
        "id": "1988_1998",
        "title": "WARS 1988 to 1998",
        "url": "https://www2.gov.bc.ca/assets/gov/driving-and-transportation/transportation-infrastructure/engineering-standards-and-guidelines/environment/wars/historical-data/1988_to_1998_wars_historic_data.xlsx",
    },
    {
        "id": "1978_1988",
        "title": "WARS 1978 to 1988",
        "url": "https://www2.gov.bc.ca/assets/gov/driving-and-transportation/transportation-infrastructure/engineering-standards-and-guidelines/environment/wars/historical-data/1978_to_1988_wars_historic_data.xlsx",
    },
]

FIELDNAMES = [
    "id",
    "accidentDate",
    "year",
    "timeOfKill",
    "nearestTown",
    "species",
    "sex",
    "age",
    "quantity",
    "latitude",
    "longitude",
    "serviceArea",
    "dataSet",
    "sourceFile",
]


def normalized_town(value):
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def is_prince_george_town(value):
    town = normalized_town(value)
    return town == "PG" or town.startswith("PRINCEGEORGE") or town.startswith("PRINCEGORGE")


def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text


def parse_number(value):
    if value is None or value == "":
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return ""
    if number.is_integer():
        return int(number)
    return number


def download(url, path):
    request = Request(url, headers={"User-Agent": "PGMaps data sync (https://github.com/) Python"})
    with urlopen(request) as response:
        path.write_bytes(response.read())


def rows_from_workbook(path, source_file):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]
    index = {header: offset for offset, header in enumerate(headers)}

    for row in rows:
        nearest_town = clean_text(row[index["Nearest.Town"]])
        if not is_prince_george_town(nearest_town):
            continue

        latitude = parse_number(row[index["Latitude"]])
        longitude = parse_number(row[index["Longitude"]])
        if latitude == "" or longitude == "":
            continue

        yield {
            "id": clean_text(row[index["ID"]]),
            "accidentDate": parse_date(row[index["Accident.Date"]]),
            "year": parse_number(row[index["Year"]]),
            "timeOfKill": clean_text(row[index["Time.of.Kill"]]),
            "nearestTown": nearest_town,
            "species": clean_text(row[index["Species"]]) or "Unknown",
            "sex": clean_text(row[index["Sex"]]),
            "age": clean_text(row[index["Age"]]),
            "quantity": parse_number(row[index["Quantity"]]) or 1,
            "latitude": latitude,
            "longitude": longitude,
            "serviceArea": parse_number(row[index["Service.Area"]]),
            "dataSet": clean_text(row[index["Data.Set"]]),
            "sourceFile": source_file["id"],
        }


def to_geojson(rows):
    return {
        "type": "FeatureCollection",
        "features": [
            {
                "type": "Feature",
                "id": row["id"],
                "geometry": {
                    "type": "Point",
                    "coordinates": [row["longitude"], row["latitude"]],
                },
                "properties": {
                    "id": row["id"],
                    "accidentDate": row["accidentDate"],
                    "year": row["year"],
                    "timeOfKill": row["timeOfKill"],
                    "nearestTown": row["nearestTown"],
                    "species": row["species"],
                    "sex": row["sex"],
                    "age": row["age"],
                    "quantity": row["quantity"],
                    "serviceArea": row["serviceArea"],
                    "dataSet": row["dataSet"],
                    "sourceFile": row["sourceFile"],
                },
            }
            for row in rows
        ],
    }


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    all_rows = []
    source_summaries = []

    with TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        for source_file in SOURCE_FILES:
            xlsx_path = tmp_path / f"{source_file['id']}.xlsx"
            print(f"Downloading {source_file['title']}")
            download(source_file["url"], xlsx_path)
            file_rows = list(rows_from_workbook(xlsx_path, source_file))
            all_rows.extend(file_rows)
            source_summaries.append({
                **source_file,
                "rows": len(file_rows),
            })
            print(f"{source_file['title']}: {len(file_rows)} Prince George rows")

    all_rows.sort(key=lambda row: (int(row["year"] or 0), row["accidentDate"], row["id"]))

    csv_path = OUTPUT_DIR / "prince_george_wildlife_accidents.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=FIELDNAMES)
        writer.writeheader()
        writer.writerows(all_rows)

    geojson_path = OUTPUT_DIR / "prince_george_wildlife_accidents.geojson"
    geojson_path.write_text(f"{json.dumps(to_geojson(all_rows), separators=(',', ':'))}\n", encoding="utf-8")

    species_counts = Counter()
    year_counts = Counter()
    total_quantity = 0
    for row in all_rows:
        quantity = int(row["quantity"] or 0)
        species_counts[row["species"]] += quantity
        year_counts[str(row["year"])] += quantity
        total_quantity += quantity

    years = sorted(int(year) for year in year_counts if year.isdigit())
    manifest = {
        "source": "BC Ministry of Transportation and Transit Wildlife Accident Reporting System",
        "sourcePage": SOURCE_PAGE,
        "sourceLicense": SOURCE_LICENSE,
        "sourceCitation": SOURCE_CITATION,
        "coverage": "Prince George nearest-town WARS records with coordinates",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "csv": "/data/wars/prince_george_wildlife_accidents.csv",
        "geojson": "/data/wars/prince_george_wildlife_accidents.geojson",
        "rows": len(all_rows),
        "totalQuantity": total_quantity,
        "yearStart": years[0] if years else None,
        "yearEnd": years[-1] if years else None,
        "species": [{"name": name, "count": count} for name, count in species_counts.most_common()],
        "years": [{"year": int(year), "count": count} for year, count in sorted(year_counts.items()) if year.isdigit()],
        "sourceFiles": source_summaries,
        "fields": FIELDNAMES,
    }
    (OUTPUT_DIR / "manifest.json").write_text(f"{json.dumps(manifest, indent=2)}\n", encoding="utf-8")
    print(f"Wrote {len(all_rows)} rows to {OUTPUT_DIR}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(exc, file=sys.stderr)
        sys.exit(1)
